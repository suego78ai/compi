"""
scrape_excel.py
----------------
엑셀 파일(.xlsx)의 대학 경쟁률 URL을 로컬 파이썬 엔진으로 고속 스크래핑하여
로컬 DB(ipsi.db)에 저장하고, data/data.json 및 GitHub Pages로 자동 배포합니다.

사용법:
    python scrape_excel.py [엑셀파일명.xlsx] [--push]
"""

import sys
import os
import json
import re
from pathlib import Path
import pandas as pd
import openpyxl

from database import SessionLocal, University, DepartmentData, engine, Base
from scraper_service import scrape_university_data
from export_data import export_to_json, push_to_github_pages

def detect_and_extract_rows(file_path):
    """
    엑셀 파일의 모든 시트를 순회하여 헤더 유무 및 열 순서에 상관없이
    (연도, 모집시기, 대학명, 무료접수, 중복지원, URL) 목록을 자동 추출합니다.
    하이퍼링크(hyperlink) URL도 함께 감지합니다.
    """
    wb = openpyxl.load_workbook(file_path, data_only=False)
    extracted = []
    seen_urls = set()
    
    for sname in wb.sheetnames:
        sheet = wb[sname]
        
        # 시트명에서 기본 모집시기 유추
        default_adm = "수시1차"
        if "수시2" in sname or "2차" in sname:
            default_adm = "수시2차"
        elif "정시" in sname:
            default_adm = "정시"
        elif "수시1" in sname or "1차" in sname:
            default_adm = "수시1차"
            
        default_year = "2026"
        m_sy = re.search(r"(202[0-9])", sname)
        if m_sy:
            default_year = m_sy.group(1)
            
        for r_idx, row in enumerate(sheet.iter_rows(values_only=False)):
            if not row:
                continue
                
            # 1. URL 찾기 (셀 텍스트 및 하이퍼링크)
            url = ""
            url_col_idx = -1
            row_strs = []
            
            for i, cell in enumerate(row):
                val = str(cell.value or "").strip()
                row_strs.append(val)
                if not url:
                    if val.startswith("http://") or val.startswith("https://"):
                        url = val
                        url_col_idx = i
                    elif cell.hyperlink and cell.hyperlink.target:
                        target = str(cell.hyperlink.target).strip()
                        if target.startswith("http://") or target.startswith("https://"):
                            url = target
                            url_col_idx = i
                            
            if not url or url in seen_urls:
                continue
            seen_urls.add(url)
                
            # 2. 연도 찾기 (4자리 숫자 202x)
            year = default_year
            for i, val in enumerate(row_strs):
                if i != url_col_idx:
                    m = re.search(r"(202[0-9])", val)
                    if m:
                        year = m.group(1)
                        break
                        
            # 3. 모집시기 찾기 (수시1차, 수시2차, 정시 등)
            adm_type = default_adm
            for i, val in enumerate(row_strs):
                if i != url_col_idx:
                    if "수시1" in val or "1차" in val:
                        adm_type = "수시1차"
                        break
                    elif "수시2" in val or "2차" in val:
                        adm_type = "수시2차"
                        break
                    elif "정시" in val:
                        adm_type = "정시"
                        break
                    elif "수시" in val:
                        adm_type = "수시"
                        break
                        
            # 4. 정원구분 기본값 '구분없음' (엑셀 서식 컬럼에서 삭제됨)
            cap_type = "구분없음"
                        
            # 5. 무료접수 ("F" 또는 "무료") 찾기
            free_apply = ""
            for i, val in enumerate(row_strs):
                if i != url_col_idx:
                    if val.strip() in ["F", "f", "무료", "Y", "y", "O", "o"]:
                        free_apply = "F"
                        break

            # 6. 중복지원 ("M" 또는 "중복") 찾기
            multi_apply = ""
            for i, val in enumerate(row_strs):
                if i != url_col_idx:
                    if val.strip() in ["M", "m", "중복", "복수", "복수지원"]:
                        multi_apply = "M"
                        break

            # 7. 대학명 찾기
            name = ""
            for i, val in enumerate(row_strs):
                if i != url_col_idx and val:
                    if val in [year, adm_type, cap_type, free_apply, multi_apply, "F", "f", "M", "m", "무료", "중복", "복수", "해당없음", "Y", "N"]:
                        continue
                    if any(h in val for h in ["대학", "대학교", "전문대학", "대"]):
                        name = val
                        break
                    elif not name and len(val) >= 2 and not val.isdigit():
                        name = val
                        
            if not name:
                for i in range(url_col_idx - 1, -1, -1):
                    if row_strs[i] and row_strs[i] not in [year, adm_type, free_apply, multi_apply]:
                        name = row_strs[i]
                        break
                        
            name = name or "대학"
            if not free_apply and (name.endswith("F") or name.endswith("(F)")):
                free_apply = "F"
            if not multi_apply and (name.endswith("M") or name.endswith("(M)") or name.endswith("[M]")):
                multi_apply = "M"

            extracted.append({
                "year": year,
                "admission_type": adm_type,
                "name": name,
                "is_free_apply": free_apply,
                "is_multi_apply": multi_apply,
                "url": url,
                "capacity_type": cap_type
            })
            
    return extracted

def process_excel_scrape(file_path, auto_push=True):
    print(f"========================================================")
    print(f" 🚀 [로컬 엑셀 스크래핑 엔진] 파일: {Path(file_path).name}")
    print(f"========================================================")
    
    if not os.path.exists(file_path):
        print(f"[오류] 파일을 찾을 수 없습니다: {file_path}")
        return False
        
    rows = detect_and_extract_rows(file_path)
    if not rows:
        print(f"[오류] 엑셀 파일에서 유효한 대학 경쟁률 URL을 찾지 못했습니다.")
        return False
        
    print(f"📋 총 {len(rows)}개 대학 경쟁률 URL 감지됨:")
    for idx, r in enumerate(rows, 1):
        f_tag = " [무료(F)]" if r.get('is_free_apply') == 'F' else ""
        m_tag = " [중복(M)]" if r.get('is_multi_apply') == 'M' else ""
        print(f"   [{idx}/{len(rows)}] {r['name']}{f_tag}{m_tag} ({r['year']} {r['admission_type']}) -> {r['url'][:50]}...")
    print("--------------------------------------------------------")
    
    db = SessionLocal()
    success_count = 0
    failed_count = 0
    
    try:
        for idx, item in enumerate(rows, 1):
            name = item["name"]
            year = item["year"]
            adm = item["admission_type"]
            cap = item["capacity_type"]
            free = item.get("is_free_apply", "")
            multi = item.get("is_multi_apply", "")
            url = item["url"]
            
            print(f"[{idx}/{len(rows)}] 🌐 {name} 스크래핑 중...", end=" ", flush=True)
            try:
                scraped_data = scrape_university_data(url)
                dept_count = len(scraped_data.get("parsed_departments", []))
                table_count = len(scraped_data.get("tables_html", []))
                
                if table_count == 0:
                    print(f"❌ 실패 (경쟁률 표 없음)")
                    failed_count += 1
                    continue
                    
                # DB 저장 또는 갱신
                existing = db.query(University).filter(
                    University.name == name,
                    University.year == year,
                    University.admission_type == adm,
                    University.capacity_type == cap
                ).first()
                
                if existing:
                    existing.url = url
                    existing.is_free_apply = free
                    existing.is_multi_apply = multi
                    existing.scraped_data = json.dumps(scraped_data)
                    db.commit()
                    univ_id = existing.id
                    # 기존 학과 데이터 삭제 후 재등록
                    db.query(DepartmentData).filter(DepartmentData.university_id == univ_id).delete()
                else:
                    new_u = University(
                        name=name,
                        year=year,
                        admission_type=adm,
                        capacity_type=cap,
                        is_free_apply=free,
                        is_multi_apply=multi,
                        url=url,
                        scraped_data=json.dumps(scraped_data)
                    )
                    db.add(new_u)
                    db.commit()
                    db.refresh(new_u)
                    univ_id = new_u.id
                    
                for d in scraped_data.get("parsed_departments", []):
                    dept = DepartmentData(
                        university_id=univ_id,
                        table_title=d.get("table_title", ""),
                        department_name=d.get("department_name", ""),
                        admission_count=str(d.get("admission_count", "")),
                        applicant_count=str(d.get("applicant_count", "")),
                        competition_ratio=str(d.get("competition_ratio", ""))
                    )
                    db.add(dept)
                db.commit()
                
                print(f"✅ 성공 ({dept_count}개 학과 추출)")
                success_count += 1
            except Exception as e:
                print(f"❌ 실패 ({e})")
                failed_count += 1
                
        print("--------------------------------------------------------")
        print(f"🎉 스크래핑 완료! 성공: {success_count}개, 실패: {failed_count}개")
        
        # JSON Export & GitHub Pages 배포
        print("\n[동기화] data.json 파일 변환 중...")
        export_to_json(db)
        
        if auto_push:
            print("\n[배포] GitHub Pages(suego78ai/ipsi)로 자동 배포 중...")
            push_to_github_pages(commit_msg=f"Update: 엑셀 스크래핑 데이터 반영 ({Path(file_path).name})")
            
        return True
    finally:
        db.close()

if __name__ == "__main__":
    target_file = None
    auto_push = True
    
    for arg in sys.argv[1:]:
        if arg in ["--no-push", "-n"]:
            auto_push = False
        elif arg.endswith(".xlsx") and os.path.exists(arg):
            target_file = arg
            
    if not target_file:
        # 현재 폴더에 있는 xlsx 파일 중 URL이 있는 파일 자동 탐색
        candidates = [f for f in os.listdir(".") if f.endswith(".xlsx") and not f.startswith("~$") and f != "template.xlsx"]
        best_candidate = None
        best_count = 0
        for c in candidates:
            try:
                extracted = detect_and_extract_rows(c)
                if len(extracted) > best_count:
                    best_count = len(extracted)
                    best_candidate = c
            except Exception:
                pass
        if best_candidate:
            target_file = best_candidate
            print(f"대상 엑셀 파일을 지정하지 않아 URL이 감지된 '{target_file}' ({best_count}개 대학)으로 자동 진행합니다.")
        elif candidates:
            target_file = candidates[0]
            print(f"파일을 지정하지 않아 '{target_file}' 파일을 대상으로 진행합니다.")
        else:
            print("사용법: python scrape_excel.py <엑셀파일명.xlsx>")
            sys.exit(1)
            
    process_excel_scrape(target_file, auto_push=auto_push)
